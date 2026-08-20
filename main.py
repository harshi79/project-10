"""
Yori Prime Cleaner Bot — super advanced .txt filtering bot.

Reads almost any .txt file, auto-detects each line's type, cleans it,
deduplicates, and lets you filter the output by many criteria:

  TYPES     : Cards, Email combos, Phone combos, Proxies, URLs, Crypto
  FILTERS   : Brand (card BIN) · Country tag · Phone code · Proxy protocol
              Proxy port · Crypto network · Domain · Sort by domain
  CLEANING  : junk removal (URLs, tg headers, # comments, blanks), dedupe
  OUTPUT    : TXT / CSV / Excel (.xlsx) with summary + per-type sheets

Upload limit: 20 MB (Telegram bot download limit).

All data used is fake/test data — this is an educational project about how
filtering works, nothing else. No account or service is ever contacted.

Owner: https://t.me/yorichiiprime
"""

import os
import re
import csv as _csv
import sqlite3
import logging
import threading
from io import BytesIO, StringIO
from datetime import datetime, timezone
from collections import defaultdict
from http.server import HTTPServer, BaseHTTPRequestHandler

import aiohttp
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)
from telegram.error import Forbidden, BadRequest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ──────────────────────────────────────────────────────────────────────

TOKEN        = os.environ.get("TELEGRAM_BOT_TOKEN", "")
OWNER_ID     = 7728424218
BRAND_URL    = "https://t.me/yorichiiprime"
BRAND_HANDLE = "@yorichiiprime"
BRAND_NAME   = "Yori Prime"
MAX_FILE_MB  = 20                          # Telegram bot download limit
MAX_BYTES    = MAX_FILE_MB * 1024 * 1024
WATERMARK    = f"\n\n— {BRAND_NAME}\n👑 Owner: {BRAND_HANDLE}\n{BRAND_URL}"
FOOTER       = f"— {BRAND_NAME}\n👑 Owner: {BRAND_URL}"
WEBHOOK_URL  = os.environ.get("WEBHOOK_URL", "")
PORT         = int(os.environ.get("PORT", 8080))
DB_PATH      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.db")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ── Queue ───────────────────────────────────────────────────────────────────────

file_queue: list[dict] = []
queue_lock = threading.Lock()

# ── Health server ───────────────────────────────────────────────────────────────
# A tiny, dependency-free HTTP server used by Render's health check and by
# UptimeRobot to keep the (free) service awake. It runs alongside the bot's
# polling loop on the same $PORT.

HEALTH_BODY = b"OK - Yori Prime Bot is alive\n"


class _Health(BaseHTTPRequestHandler):
    def _respond(self):
        path = self.path.split("?", 1)[0]
        if path in ("/", "/health", "/healthz"):
            self.send_response(200)
            self.send_header("Content-Type", "text/plain")
            self.send_header("Content-Length", str(len(HEALTH_BODY)))
            self.end_headers()
            if self.command != "HEAD":
                self.wfile.write(HEALTH_BODY)
        else:
            self.send_response(404)
            self.send_header("Content-Length", "0")
            self.end_headers()

    def do_GET(self):
        self._respond()

    def do_HEAD(self):
        self._respond()

    def log_message(self, *a):
        pass


def _start_health():
    try:
        server = HTTPServer(("0.0.0.0", PORT), _Health)
        log.info("Health server listening on http://0.0.0.0:%d/health", PORT)
        server.serve_forever()
    except OSError as e:
        log.warning("Health server could not bind port %d: %s", PORT, e)

# ── Database ────────────────────────────────────────────────────────────────────

def _conn() -> sqlite3.Connection:
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def init_db() -> None:
    with _conn() as c:
        c.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id        INTEGER PRIMARY KEY,
                name      TEXT    DEFAULT '',
                username  TEXT    DEFAULT '',
                files     INTEGER DEFAULT 0,
                lines     INTEGER DEFAULT 0,
                cards     INTEGER DEFAULT 0,
                combos    INTEGER DEFAULT 0,
                proxies   INTEGER DEFAULT 0,
                urls      INTEGER DEFAULT 0,
                cryptos   INTEGER DEFAULT 0,
                last_seen TEXT    DEFAULT ''
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS queue_history (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                uid      INTEGER,
                filename TEXT,
                status   TEXT,
                ts       TEXT
            )
        """)
        # Migrate DBs created before the extra columns existed
        cols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
        for col in ("proxies", "urls", "cryptos"):
            if col not in cols:
                c.execute(f"ALTER TABLE users ADD COLUMN {col} INTEGER DEFAULT 0")


def upsert_user(uid: int, name: str, username: str, counts: dict) -> None:
    n_cards   = counts.get("cards", 0)
    n_combos  = counts.get("emails", 0) + counts.get("phones", 0)
    n_proxies = counts.get("proxies", 0)
    n_urls    = counts.get("urls", 0)
    n_cryptos = counts.get("cryptos", 0)
    total = n_cards + n_combos + n_proxies + n_urls + n_cryptos
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    with _conn() as c:
        c.execute("""
            INSERT INTO users (id, name, username, files, lines, cards, combos, proxies, urls, cryptos, last_seen)
            VALUES (?,?,?,1,?,?,?,?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET
                name      = excluded.name,
                username  = excluded.username,
                files     = files  + 1,
                lines     = lines  + excluded.lines,
                cards     = cards  + excluded.cards,
                combos    = combos + excluded.combos,
                proxies   = proxies + excluded.proxies,
                urls      = urls   + excluded.urls,
                cryptos   = cryptos + excluded.cryptos,
                last_seen = excluded.last_seen
        """, (uid, name, username, total, n_cards, n_combos, n_proxies, n_urls, n_cryptos, ts))


def get_user(uid: int):
    with _conn() as c:
        return c.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()


def get_all_user_ids() -> list[int]:
    with _conn() as c:
        return [r[0] for r in c.execute("SELECT id FROM users").fetchall()]


def get_global_stats():
    with _conn() as c:
        total_users = c.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        total_files = c.execute("SELECT COALESCE(SUM(files),0) FROM users").fetchone()[0]
        total_lines = c.execute("SELECT COALESCE(SUM(lines),0) FROM users").fetchone()[0]
        top5 = c.execute(
            "SELECT name, username, lines FROM users ORDER BY lines DESC LIMIT 5"
        ).fetchall()
    return total_users, total_files, total_lines, top5


def log_queue(uid: int, filename: str, status: str) -> None:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    with _conn() as c:
        c.execute("INSERT INTO queue_history (uid, filename, status, ts) VALUES (?,?,?,?)",
                  (uid, filename, status, ts))


def get_queue_history(uid: int, limit: int = 10):
    with _conn() as c:
        return c.execute(
            "SELECT filename, status, ts FROM queue_history WHERE uid=? ORDER BY id DESC LIMIT ?",
            (uid, limit)
        ).fetchall()


# ── Smart Line Analyser ─────────────────────────────────────────────────────────

CARD_RE = re.compile(
    r"^(\d{13,19})[\s|:;]+(\d{1,2})[\s|:;]+(\d{2,4})[\s|:;]+(\d{3,4})"
    r"(?:[\s]*[\u2014\u2013\-]+.*)?$"
)
EMAIL_RE       = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]{2,}$")
EMAIL_SEP_RE   = re.compile(r"^([^\s@:;|]+@[^\s@:;|]+\.[^\s@:;|]{2,})[:;|](.+)$")
EMAIL_SPACE_RE = re.compile(r"^([^\s@]+@[^\s@]+\.[^\s@]{2,})\s+(\S+)$")
PHONE_SEP_RE   = re.compile(r"^(\+\d{7,15}|\d{7,12})[:;|](.+)$")
PHONE_SPACE_RE = re.compile(r"^(\+\d{7,15})\s+(\S+)$")
URL_RE         = re.compile(r"^https?://\S+$", re.I)
TG_HEAD_RE     = re.compile(r"^.{1,80},\s*\[\d{1,2}/\d{1,2}/\d{4}")
# Telegram-only links are junk; real http(s) links are a data type now.
JUNK_RE        = re.compile(r"^(tg://|t\.me/)", re.I)

ETH_RE  = re.compile(r"^0x[a-fA-F0-9]{40}$")
BTC_BECH32_RE = re.compile(r"^bc1[a-zA-Z0-9]{25,59}$")
BTC_BASE58_RE = re.compile(r"^[13][1-9A-HJ-NP-Za-km-z]{25,34}$")

# Country tag at the end of a line: "… — 🇦🇪 AE"  (flag optional)
COUNTRY_RE = re.compile(
    r"[\u2014\u2013\-]\s*(?:[\U0001F1E6-\U0001F1FF]{2}\s*)?([A-Za-z]{2})\s*$"
)

# Common ITU country calling codes (E.164), longest-prefix matched.
COUNTRY_CODES = {
    "1", "7", "20", "27", "30", "31", "32", "33", "34", "36", "39", "40",
    "41", "43", "44", "45", "46", "47", "48", "49", "51", "52", "53", "54",
    "55", "56", "57", "58", "60", "61", "62", "63", "64", "65", "66", "81",
    "82", "84", "86", "90", "91", "92", "93", "94", "95", "98", "211", "212",
    "213", "216", "218", "220", "221", "222", "223", "224", "225", "226",
    "227", "228", "229", "230", "231", "232", "233", "234", "235", "236",
    "237", "238", "239", "240", "241", "242", "243", "244", "245", "246",
    "247", "248", "249", "250", "251", "252", "253", "254", "255", "256",
    "257", "258", "260", "261", "262", "263", "264", "265", "266", "267",
    "268", "269", "290", "291", "297", "298", "299", "350", "351", "352",
    "353", "354", "355", "356", "357", "358", "359", "370", "371", "372",
    "373", "374", "375", "376", "377", "378", "380", "381", "382", "383",
    "385", "386", "387", "389", "420", "421", "423", "500", "501", "502",
    "503", "504", "505", "506", "507", "508", "509", "590", "591", "592",
    "593", "594", "595", "596", "597", "598", "599", "670", "672", "673",
    "674", "675", "676", "677", "678", "679", "680", "681", "682", "683",
    "685", "686", "687", "688", "689", "690", "691", "692", "850", "852",
    "853", "855", "856", "880", "886", "960", "961", "962", "963", "964",
    "965", "966", "967", "968", "970", "971", "972", "973", "974", "975",
    "976", "977", "992", "993", "994", "995", "996", "998",
}
CODE_KEYS = sorted(COUNTRY_CODES, key=lambda c: (-len(c), -int(c)))

# Brand detection from BIN (first digits) — used for FILTERING.
BRAND_RULES = [
    ("Visa",             lambda n: n[0] == "4"),
    ("Mastercard",       lambda n: len(n) >= 2 and 51 <= int(n[:2]) <= 55),
    ("Mastercard",       lambda n: len(n) >= 4 and 2221 <= int(n[:4]) <= 2720),
    ("American Express", lambda n: n.startswith(("34", "37"))),
    ("Discover",         lambda n: len(n) >= 6 and 622126 <= int(n[:6]) <= 622925),
    ("Discover",         lambda n: n.startswith(("6011", "65"))),
    ("JCB",              lambda n: len(n) >= 4 and 3528 <= int(n[:4]) <= 3589),
    ("UnionPay",         lambda n: n.startswith("62")),
    ("Diners Club",      lambda n: n.startswith(("300", "301", "302", "303", "304", "305", "36", "38"))),
    ("Maestro",          lambda n: n.startswith(("50", "56", "57", "58", "63"))),
]


def card_brand(number: str) -> str:
    n = "".join(c for c in number if c.isdigit())
    for brand, rule in BRAND_RULES:
        if n and rule(n):
            return brand
    return "Unknown"


def extract_country(raw: str):
    m = COUNTRY_RE.search(raw)
    if m:
        return m.group(1).upper(), raw[:m.start()].rstrip()
    return None, raw


def phone_code(phone: str):
    m = re.match(r"^\+(\d+)", phone.strip())
    if not m:
        return None
    digits = m.group(1)
    for code in CODE_KEYS:
        if digits.startswith(code):
            return f"+{code}"
    return None


def _valid_ipv4(host: str) -> bool:
    octets = host.split(".")
    return (len(octets) == 4
            and all(o.isdigit() and 0 <= int(o) <= 255 for o in octets))


def parse_proxy(t: str):
    """Parse proxy line into a dict, or None.

    Supports: host:port, scheme://host:port, scheme://user:pass@host:port,
    host:port:user:pass, host:port|user|pass.
    """
    s = t.strip()
    protocol = None
    m = re.match(r"^(http|https|socks4|socks5)://(.+)$", s, re.I)
    if m:
        protocol = m.group(1).lower()
        s = m.group(2)
    if "/" in s:          # has a path → it's a URL, not a proxy
        return None

    user = pw = None
    if "@" in s:
        cred, hostport = s.rsplit("@", 1)
        if ":" in cred:
            user, pw = cred.split(":", 1)
        else:
            return None
        s = hostport

    parts = re.split(r"[:\|]", s)
    if len(parts) < 2:
        return None
    host, port_s = parts[0], parts[1]
    if not host or not port_s.isdigit():
        return None
    port = int(port_s)
    if not (1 <= port <= 65535):
        return None

    extra = parts[2:]
    if extra:
        if user is None:
            user = extra[0]
        if len(extra) >= 2 and pw is None:
            pw = extra[1]

    if _valid_ipv4(host):
        host_type = "ip"
    elif re.fullmatch(r"[a-zA-Z0-9.\-]+", host) and "." in host:
        host_type = "host"
    else:
        return None

    return {
        "host": host, "port": port,
        "protocol": protocol or "unknown",
        "user": user, "pw": pw, "host_type": host_type,
    }


def analyse_line(raw: str):
    """Classify one line. Returns a tuple tagged with the type, or None."""
    t = raw.strip()
    if not t or t.startswith("#"):
        return None
    if JUNK_RE.match(t) or TG_HEAD_RE.match(t):
        return None

    country, t = extract_country(t)
    t = t.strip()
    if not t:
        return None

    m = CARD_RE.match(t)
    if m:
        num, mm, yy, cvv = m.groups()
        return ("card", num, mm, yy, cvv, country)

    m = EMAIL_SEP_RE.match(t)
    if m:
        email, pw = m.group(1).strip(), m.group(2).strip()
        if EMAIL_RE.match(email) and pw:
            return ("combo", email, pw, country)

    if "@" in t:
        m = EMAIL_SPACE_RE.match(t)
        if m:
            email, pw = m.group(1).strip(), m.group(2).strip()
            if EMAIL_RE.match(email) and pw:
                return ("combo", email, pw, country)

    m = PHONE_SEP_RE.match(t)
    if m:
        phone, pw = m.group(1).strip(), m.group(2).strip()
        if pw:
            return ("phone", phone, pw, country, phone_code(phone))

    m = PHONE_SPACE_RE.match(t)
    if m:
        phone, pw = m.group(1).strip(), m.group(2).strip()
        if pw:
            return ("phone", phone, pw, country, phone_code(phone))

    p = parse_proxy(t)
    if p:
        p["country"] = country
        return ("proxy", p)

    if URL_RE.match(t):
        host = re.sub(r"^https?://", "", t, flags=re.I).split("/")[0].split(":")[0].lower()
        return ("url", t, host, country)

    if ETH_RE.match(t):
        return ("crypto", t, "ETH", country)
    if BTC_BECH32_RE.match(t) or BTC_BASE58_RE.match(t):
        return ("crypto", t, "BTC", country)

    return None


def analyse_file(content: str) -> dict:
    buckets = {
        "cards": [], "combos": [], "phones": [],
        "proxies": [], "urls": [], "cryptos": [],
    }
    seen = {k: set() for k in buckets}
    skipped = 0
    total_nonempty = sum(
        1 for l in content.splitlines()
        if l.strip() and not l.lstrip().startswith("#")
    )
    removed_links = []   # collect t.me links removed from output

    for raw in content.splitlines():
        r = analyse_line(raw)
        if r is None:
            if raw.strip() and not raw.lstrip().startswith("#"):
                skipped += 1
            continue

        kind = r[0]

        if kind == "card":
            _, num, mm, yy, cvv, country = r
            value = f"{num}|{mm}|{yy}|{cvv}"
            if value in seen["cards"]:
                skipped += 1
                continue
            seen["cards"].add(value)
            buckets["cards"].append({
                "value": value, "num": num, "mm": mm, "yy": yy, "cvv": cvv,
                "brand": card_brand(num), "country": country,
            })

        elif kind == "combo":
            _, email, pw, country = r
            key = f"{email.lower()}:::{pw}"
            if key in seen["combos"]:
                skipped += 1
                continue
            seen["combos"].add(key)
            buckets["combos"].append({
                "email": email, "pw": pw,
                "domain": email.rsplit("@", 1)[-1].lower(), "country": country,
            })

        elif kind == "phone":
            _, phone, pw, country, code = r
            key = f"{phone}:::{pw}"
            if key in seen["phones"]:
                skipped += 1
                continue
            seen["phones"].add(key)
            buckets["phones"].append({
                "phone": phone, "pw": pw, "code": code, "country": country,
            })

        elif kind == "proxy":
            p = r[1]
            key = f"{p['host']}:{p['port']}:{p['user']}:{p['pw']}:{p['protocol']}"
            if key in seen["proxies"]:
                skipped += 1
                continue
            seen["proxies"].add(key)
            buckets["proxies"].append(p)

        elif kind == "url":
            _, url, host, country = r
            # Remove t.me / https://t.me links from output
            if "t.me/" in url or "https://t.me/" in url:
                removed_links.append(url)
                skipped += 1
                continue
            if url in seen["urls"]:
                skipped += 1
                continue
            seen["urls"].add(url)
            buckets["urls"].append({"url": url, "domain": host, "country": country})

        elif kind == "crypto":
            _, addr, net, country = r
            if addr in seen["cryptos"]:
                skipped += 1
                continue
            seen["cryptos"].add(addr)
            buckets["cryptos"].append({"address": addr, "network": net, "country": country})

    return {**buckets, "skipped": skipped, "total": total_nonempty, "removed_links": removed_links}


# ── Filter helpers ──────────────────────────────────────────────────────────────

def _domain_stats(combos: list, urls: list) -> dict:
    counts: dict[str, int] = defaultdict(int)
    for c in combos:
        counts[c["domain"]] += 1
    for u in urls:
        counts[u["domain"]] += 1
    return counts


def get_domains(combos: list, urls: list) -> list[str]:
    counts = _domain_stats(combos, urls)
    return sorted(counts.keys(), key=lambda d: (-counts[d], d))


def sort_combos_by_domain(combos: list) -> list:
    return sorted(combos, key=lambda c: (c["domain"], c["email"].lower()))


def sort_urls_by_domain(urls: list) -> list:
    return sorted(urls, key=lambda u: (u["domain"], u["url"].lower()))


def filter_by_domain(combos: list, urls: list, domain: str):
    if not domain:
        return combos, urls
    d = domain.lower()
    return ([c for c in combos if c["domain"] == d],
            [u for u in urls if u["domain"] == d])


def get_available_brands(cards: list) -> list:
    return sorted({c["brand"] for c in cards})


def get_available_countries(*groups) -> list:
    codes = set()
    for g in groups:
        for e in g:
            if e.get("country"):
                codes.add(e["country"])
    return sorted(codes)


def get_available_codes(phones: list) -> list:
    return sorted({c["code"] for c in phones if c.get("code")})


def get_available_protocols(proxies: list) -> list:
    return sorted({p["protocol"] for p in proxies if p.get("protocol") != "unknown"})


def get_available_ports(proxies: list) -> list:
    return sorted({p["port"] for p in proxies})


def get_available_networks(cryptos: list) -> list:
    return sorted({c["network"] for c in cryptos})


def next_choice(current, options: list):
    seq = [None] + list(options)
    try:
        i = seq.index(current)
    except ValueError:
        i = 0
    return seq[(i + 1) % len(seq)]


# ── Output Builders ─────────────────────────────────────────────────────────────

def proxy_str(p: dict) -> str:
    auth = f"{p['user']}:{p['pw']}@" if p.get("user") else ""
    if p.get("protocol") and p["protocol"] != "unknown":
        return f"{p['protocol']}://{auth}{p['host']}:{p['port']}"
    return f"{auth}{p['host']}:{p['port']}"


def build_txt(cards, combos, phones, proxies, urls, cryptos) -> str:
    groups = [
        ("CARDS", cards, lambda c: c["value"]),
        ("COMBOS", combos, lambda c: f"{c['email']}   {c['pw']}"),
        ("PHONES", phones, lambda c: f"{c['phone']}   {c['pw']}"),
        ("PROXIES", proxies, proxy_str),
        ("URLS", urls, lambda u: u["url"]),
        ("CRYPTO", cryptos, lambda c: c["address"]),
    ]
    present = [(n, g, f) for n, g, f in groups if g]
    if not present:
        return ""
    if len(present) == 1:
        name, g, fmt = present[0]
        return "\n".join(fmt(x) for x in g) + WATERMARK
    parts: list[str] = []
    for name, g, fmt in present:
        parts += [f"━━━ {name} ({len(g)}) ━━━", *[fmt(x) for x in g], ""]
    return "\n".join(parts).rstrip("\n") + WATERMARK


def build_csv(cards, combos, phones, proxies, urls, cryptos) -> str:
    rows: list[list] = []
    if cards:
        rows.append(["Type", "Number", "Month", "Year", "CVV", "Brand", "Country"])
        for c in cards:
            rows.append(["card", c["num"], c["mm"], c["yy"], c["cvv"],
                         c["brand"], c.get("country") or ""])
    if combos:
        rows.append(["Type", "Email", "Password", "Domain", "Country"])
        for c in combos:
            rows.append(["combo", c["email"], c["pw"], c["domain"], c.get("country") or ""])
    if phones:
        rows.append(["Type", "Phone", "Password", "Code", "Country"])
        for c in phones:
            rows.append(["phone", c["phone"], c["pw"], c.get("code") or "", c.get("country") or ""])
    if proxies:
        rows.append(["Type", "Host", "Port", "Protocol", "User", "Pass", "HostType", "Country"])
        for p in proxies:
            rows.append(["proxy", p["host"], p["port"], p["protocol"],
                         p["user"] or "", p["pw"] or "", p["host_type"],
                         p.get("country") or ""])
    if urls:
        rows.append(["Type", "URL", "Domain", "Country"])
        for u in urls:
            rows.append(["url", u["url"], u["domain"], u.get("country") or ""])
    if cryptos:
        rows.append(["Type", "Address", "Network", "Country"])
        for c in cryptos:
            rows.append(["crypto", c["address"], c["network"], c.get("country") or ""])
    buf = StringIO()
    _csv.writer(buf).writerows(rows)
    return buf.getvalue()


def _style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center")


def _write_table(ws, headers: list, rows: list) -> None:
    ws.append(headers)
    for cell in ws[1]:
        _style_header(cell)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        letter = col[0].column_letter
        m = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(m + 2, 8), 42)


def _write_summary_sheet(ws, groups: dict) -> None:
    ws.append([f"{BRAND_NAME} — Filtered Output"])
    ws.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])
    ws.append(["Type", "Total"])
    for cell in ws[4]:
        _style_header(cell)
    for name, entries in groups.items():
        ws.append([name, len(entries)])
    ws.append([])
    ws.append(["Owner", BRAND_URL])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40


def build_xlsx(cards, combos, phones, proxies, urls, cryptos) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _write_summary_sheet(ws, {
        "Cards": cards, "Emails": combos, "Phones": phones,
        "Proxies": proxies, "URLs": urls, "Crypto": cryptos,
    })

    if cards:
        s = wb.create_sheet("Cards")
        _write_table(s,
                     ["Number", "Month", "Year", "CVV", "Brand", "Country"],
                     [[c["num"], c["mm"], c["yy"], c["cvv"], c["brand"],
                       c.get("country") or ""] for c in cards])
    if combos:
        s = wb.create_sheet("Emails")
        _write_table(s,
                     ["Email", "Password", "Domain", "Country"],
                     [[c["email"], c["pw"], c["domain"], c.get("country") or ""] for c in combos])
    if phones:
        s = wb.create_sheet("Phones")
        _write_table(s,
                     ["Phone", "Password", "Code", "Country"],
                     [[c["phone"], c["pw"], c.get("code") or "", c.get("country") or ""] for c in phones])
    if proxies:
        s = wb.create_sheet("Proxies")
        _write_table(s,
                     ["Host", "Port", "Protocol", "User", "Pass", "HostType", "Country"],
                     [[p["host"], p["port"], p["protocol"], p["user"] or "", p["pw"] or "",
                       p["host_type"], p.get("country") or ""] for p in proxies])
    if urls:
        s = wb.create_sheet("URLs")
        _write_table(s,
                     ["URL", "Domain", "Country"],
                     [[u["url"], u["domain"], u.get("country") or ""] for u in urls])
    if cryptos:
        s = wb.create_sheet("Crypto")
        _write_table(s,
                     ["Address", "Network", "Country"],
                     [[c["address"], c["network"], c.get("country") or ""] for c in cryptos])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Keyboards ───────────────────────────────────────────────────────────────────

OWNER_BUTTON = InlineKeyboardButton("👑 Owner", url=BRAND_URL)

MAIN_KB = ReplyKeyboardMarkup(
    [["📊 My Stats", "ℹ️ Help"], ["🏷️ About"]],
    resize_keyboard=True, is_persistent=True,
)


def _chunk(row: list, size: int = 4) -> list:
    return [row[i:i + size] for i in range(0, len(row), size)]


def type_ikb(cards, combos, phones, proxies, urls, cryptos, uid: int, fmt: str = "txt",
             selected_types: set | None = None, domain: str | None = None,
             sort: bool = False, brand: str | None = None, country: str | None = None,
             code: str | None = None, protocol: str | None = None,
             port: int | None = None, network: str | None = None) -> InlineKeyboardMarkup:
    rows = []
    sel = selected_types or set()

    # Row 1: Format buttons (success style)
    rows.append([
        InlineKeyboardButton(f"📄 TXT {'✅' if fmt == 'txt' else ''}", callback_data=f"fmt:txt:{uid}", style="success"),
        InlineKeyboardButton(f"📊 CSV {'✅' if fmt == 'csv' else ''}", callback_data=f"fmt:csv:{uid}", style="success"),
        InlineKeyboardButton(f"📈 Excel {'✅' if fmt == 'xlsx' else ''}", callback_data=f"fmt:xlsx:{uid}", style="success"),
    ])

    # Row 2: Type selection
    type_row = []
    for key, label, cb in (
        ("cards", "💳 Cards", f"sel:cards:{uid}"),
        ("combos", "🔑 Emails", f"sel:combos:{uid}"),
        ("phones", "📱 Phones", f"sel:phones:{uid}"),
        ("proxies", "🌐 Proxies", f"sel:proxies:{uid}"),
        ("urls", "🔗 URLs", f"sel:urls:{uid}"),
        ("cryptos", "💰 Crypto", f"sel:cryptos:{uid}"),
    ):
        present = bool({"cards": cards, "combos": combos, "phones": phones,
                        "proxies": proxies, "urls": urls, "cryptos": cryptos}[key])
        if present:
            type_row.append(InlineKeyboardButton(
                f"{label} {'✅' if key in sel else ''}", callback_data=cb))
    if len(type_row) > 1:
        type_row.append(InlineKeyboardButton(
            f"🔀 All {'✅' if not sel else ''}", callback_data=f"sel:all:{uid}"))
    if type_row:
        rows.extend(_chunk(type_row))

    # Row 3: Advanced filters (cycle buttons)
    filter_row = []
    if cards:
        filter_row.append(InlineKeyboardButton(
            f"💳 Brand: {brand or 'All'}", callback_data=f"brand:{uid}"))
    countries = get_available_countries(cards, combos, phones, proxies, urls, cryptos)
    if countries:
        filter_row.append(InlineKeyboardButton(
            f"🌍 Country: {country or 'All'}", callback_data=f"cty:{uid}"))
    if phones:
        filter_row.append(InlineKeyboardButton(
            f"📞 Code: {code or 'All'}", callback_data=f"code:{uid}"))
    if proxies:
        filter_row.append(InlineKeyboardButton(
            f"🌐 Protocol: {protocol or 'All'}", callback_data=f"proto:{uid}"))
        filter_row.append(InlineKeyboardButton(
            f"🔌 Port: {port or 'All'}", callback_data=f"port:{uid}"))
    if cryptos:
        filter_row.append(InlineKeyboardButton(
            f"💰 Network: {network or 'All'}", callback_data=f"net:{uid}"))
    if filter_row:
        rows.extend(_chunk(filter_row))

    # Row 4: Domain filter + sort (emails + urls)
    if combos or urls:
        domains = get_domains(combos, urls)[:5]
        if domains:
            rows.append([InlineKeyboardButton(
                f"📧 {d} {'✅' if domain == d else ''}",
                callback_data=f"seld:{d}:{uid}") for d in domains])
        rows.append([InlineKeyboardButton(
            f"🔤 Sort by domain {'✅' if sort else ''}",
            callback_data=f"sels:domain:{uid}")])

    # Row 5: Generate (primary) and Skip (danger) buttons
    rows.append([
        InlineKeyboardButton("🚀 GENERATE", callback_data=f"gen:{fmt}:{uid}", style="primary"),
        InlineKeyboardButton("⏭️ Skip", callback_data=f"skip:{uid}", style="danger"),
    ])

    # Row 6: Owner button
    rows.append([OWNER_BUTTON])

    return InlineKeyboardMarkup(rows)


def result_ikb(uid: int) -> InlineKeyboardMarkup:
    rows = [[
        InlineKeyboardButton("📊 My Stats", callback_data=f"stats:{uid}"),
        InlineKeyboardButton("ℹ️ Help", callback_data="help"),
    ]]
    if uid == OWNER_ID:
        rows.append([InlineKeyboardButton("👑 Global Stats", callback_data="owner:stats")])
    rows.append([OWNER_BUTTON])
    return InlineKeyboardMarkup(rows)


# ── Text helpers ────────────────────────────────────────────────────────────────

HELP_TEXT = (
    "<b>ℹ️ How to use</b>\n"
    "──────────────────\n"
    "Send almost any <b>.txt</b> file (up to 20 MB). Every line is detected,\n"
    "cleaned and deduplicated automatically.\n\n"
    "🔍 <b>Supported types</b>\n"
    "  💳 Cards — <code>4111111111111111|05|33|496 — 🇦🇪 AE</code>\n"
    "  🔑 Emails — <code>user@gmail.com:Password1</code>\n"
    "  📱 Phones — <code>+919876543210:Password1</code>\n"
    "  🌐 Proxies — <code>1.2.3.4:8080</code> · <code>socks5://u:p@1.2.3.4:1080</code>\n"
    "  🔗 URLs — <code>https://example.com/page</code>\n"
    "  💰 Crypto — BTC / ETH addresses\n\n"
    "🧰 <b>Filters (tap to cycle)</b>\n"
    "  💳 Brand · 🌍 Country · 📞 Code · 🌐 Protocol · 🔌 Port · 💰 Network · 📧 Domain\n\n"
    "🗑️ Junk (tg links, headers, comments) is removed automatically.\n"
    "📊 Output: TXT / CSV / Excel\n\n"
    f"👑 Owner: {BRAND_URL}\n"
    f"<i>— {BRAND_NAME}</i>"
)

ABOUT_TEXT = (
    f"<b>🏷️ {BRAND_NAME} Filter Bot</b>\n"
    "──────────────────\n"
    "⚡ Instant .txt file analysis (up to 20 MB)\n"
    "🧠 Auto-detect cards / emails / phones / proxies / URLs / crypto\n"
    "💳 Filter by card brand (BIN)\n"
    "🌍 Filter by country tag\n"
    "📞 Filter by phone country code\n"
    "🌐 Filter by proxy protocol\n"
    "🔌 Filter by proxy port\n"
    "💰 Filter by crypto network\n"
    "📧 Filter by domain + sort\n"
    "🗑️ Automatic junk removal\n"
    "🔑 Deduplication built in\n"
    "📊 TXT / CSV / Excel output\n"
    "💧 Auto-watermark on every output\n"
    "📊 Per-user stats (SQLite)\n\n"
    f"👑 Owner: {BRAND_URL}\n"
    f"<i>— {BRAND_NAME}</i>"
)


def user_stats_text(row) -> str:
    return (
        f"<b>📊 Your Stats</b>\n"
        f"──────────────────\n"
        f"👤 {row['name'] or 'Unknown'}  <i>{row['username']}</i>\n\n"
        f"📁 Files processed  <b>{row['files']:,}</b>\n"
        f"📝 Total lines      <b>{row['lines']:,}</b>\n"
        f"💳 Card lines       <b>{row['cards']:,}</b>\n"
        f"🔑 Combo lines      <b>{row['combos']:,}</b>\n"
        f"🌐 Proxy lines      <b>{row['proxies']:,}</b>\n"
        f"🔗 URL lines        <b>{row['urls']:,}</b>\n"
        f"💰 Crypto lines     <b>{row['cryptos']:,}</b>\n\n"
        f"🕒 <i>{row['last_seen']}</i>\n\n"
        f"👑 Owner: {BRAND_URL}\n"
        f"<i>— {BRAND_NAME}</i>"
    )


def global_stats_text() -> str:
    total_users, total_files, total_lines, top5 = get_global_stats()
    top_str = "\n".join(
        f"  {i + 1}. {r['name'] or r['username']} — {r['lines']:,} lines"
        for i, r in enumerate(top5)
    ) or "  No data yet."
    return (
        f"<b>👑 Global Stats</b>\n"
        f"──────────────────\n"
        f"👥 Total users   <b>{total_users:,}</b>\n"
        f"📁 Total files   <b>{total_files:,}</b>\n"
        f"📝 Total lines   <b>{total_lines:,}</b>\n\n"
        f"🏆 <b>Top 5</b>\n{top_str}\n\n"
        f"👑 Owner: {BRAND_URL}\n"
        f"<i>— {BRAND_NAME}</i>"
    )


# ── Handlers ────────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    name = update.effective_user.first_name or "there"
    await update.message.reply_text(
        f"⚡ <b>{BRAND_NAME} Filter Bot</b>\n"
        f"──────────────────────\n\n"
        f"Welcome, <b>{name}</b>.\n\n"
        f"Drop a <b>.txt</b> file (up to 20 MB) and I will:\n"
        f"  🧠 Auto-detect every line's type\n"
        f"  🗑️ Remove junk (tg links, headers, comments)\n"
        f"  🔑 Deduplicate entries\n"
        f"  💳🌍📞🌐🔌💰📧 Apply advanced filters\n"
        f"  📊 Export TXT / CSV / Excel\n\n"
        f"No commands needed — just send the file.\n\n"
        f"👑 Owner: {BRAND_URL}",
        parse_mode="HTML", reply_markup=MAIN_KB,
    )


async def cmd_stats(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != OWNER_ID:
        await update.message.reply_text("⛔ Owner only.")
        return
    await update.message.reply_text(global_stats_text(), parse_mode="HTML")


async def cmd_broadcast(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != OWNER_ID:
        await update.message.reply_text("⛔ Owner only.")
        return
    if not ctx.args:
        await update.message.reply_text(
            "Usage: /broadcast <message>\n\nBlocked users are automatically skipped."
        )
        return

    text = " ".join(ctx.args)
    users = get_all_user_ids()
    sent = blocked = failed = 0

    status_msg = await update.message.reply_text(
        f"📣 Broadcasting to {len(users):,} users…"
    )

    for uid in users:
        try:
            await ctx.bot.send_message(uid, text, parse_mode="HTML")
            sent += 1
        except (Forbidden, BadRequest):
            blocked += 1
        except Exception as e:
            log.warning("Broadcast error uid=%s: %s", uid, e)
            failed += 1

    await status_msg.edit_text(
        f"📣 <b>Broadcast complete</b>\n"
        f"──────────────────\n"
        f"✅ Sent:            <b>{sent:,}</b>\n"
        f"🚫 Blocked/skipped: <b>{blocked:,}</b>\n"
        f"❌ Other errors:    <b>{failed:,}</b>",
        parse_mode="HTML",
    )


async def cmd_queue(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    history = get_queue_history(uid)
    if not history:
        await update.message.reply_text("📭 No queue history yet.")
        return
    lines = ["<b>📋 Recent Queue</b>\n──────────────────"]
    for fn, status, ts in history:
        icon = "✅" if status == "done" else "⏳" if status == "queued" else "❌"
        lines.append(f"{icon} <code>{fn[:30]}</code> — {status} — {ts}")
    await update.message.reply_text("\n".join(lines), parse_mode="HTML")


# ── File processing ─────────────────────────────────────────────────────────────

async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    doc = update.message.document
    uid = update.effective_user.id
    user = update.effective_user

    if not doc.file_name or not doc.file_name.lower().endswith(".txt"):
        await update.message.reply_text(
            "❌ <b>Only .txt files accepted.</b>\n\n"
            "Send a plain text file — I'll take care of the rest.",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    if doc.file_size and doc.file_size > MAX_BYTES:
        mb = round(doc.file_size / 1024 / 1024, 1)
        await update.message.reply_text(
            f"❌ <b>File too large.</b>\n\n"
            f"📁 {doc.file_name}: <b>{mb} MB</b>\n"
            f"⛔ Maximum allowed: <b>{MAX_FILE_MB} MB</b> (Telegram bot download limit).\n\n"
            f"Split the file into smaller parts and send again.",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )
        return

    with queue_lock:
        user_queued = sum(1 for q in file_queue if q["uid"] == uid)
    if user_queued >= 3:
        await update.message.reply_text(
            "⏳ <b>Queue full.</b> You have 3 files pending.\n"
            "Wait for one to finish before sending more.",
            parse_mode="HTML",
        )
        return

    thinking = await update.message.reply_text(
        "🧠 <b>Analysing file…</b>", parse_mode="HTML"
    )

    try:
        tg_file = await ctx.bot.get_file(doc.file_id)
        async with aiohttp.ClientSession() as session:
            async with session.get(tg_file.file_path) as resp:
                data = await resp.read()

        if len(data) > MAX_BYTES:
            await thinking.delete()
            await update.message.reply_text(
                f"❌ <b>File too large</b> — over {MAX_FILE_MB} MB after download.\n"
                f"Split it into smaller parts.",
                parse_mode="HTML", reply_markup=MAIN_KB,
            )
            return

        content = data.decode("utf-8", errors="replace")

        result = analyse_file(content)
        cards, combos, phones = result["cards"], result["combos"], result["phones"]
        proxies, urls, cryptos = result["proxies"], result["urls"], result["cryptos"]
        skipped, total = result["skipped"], result["total"]
        removed_links = result.get("removed_links", [])

        await thinking.delete()

        # If there are removed links, send them in a separate message
        if removed_links:
            link_list = "\n".join(f"• {link}" for link in removed_links[:20])  # limit to 20 to avoid spam
            if len(removed_links) > 20:
                link_list += f"\n… and {len(removed_links)-20} more."
            await update.message.reply_text(
                f"🚫 <b>Removed t.me links from output:</b>\n\n{link_list}\n\n"
                f"These links were excluded from the final file.",
                parse_mode="HTML",
            )

        if not any([cards, combos, phones, proxies, urls, cryptos]):
            await update.message.reply_text(
                f"⚠️ <b>Nothing recognised</b> in <code>{doc.file_name}</code>.\n\n"
                f"Supported: cards, email combos, phone combos, proxies, URLs, crypto.",
                parse_mode="HTML", reply_markup=MAIN_KB,
            )
            return

        base = doc.file_name.rsplit(".txt", 1)[0].rsplit(".TXT", 1)[0]

        full_name = " ".join(filter(None, [user.first_name, user.last_name]))
        uname = f"@{user.username}" if user.username else "—"
        with queue_lock:
            file_queue.append({
                "uid": uid,
                "filename": doc.file_name,
                "base": base,
                "cards": cards, "combos": combos, "phones": phones,
                "proxies": proxies, "urls": urls, "cryptos": cryptos,
                "skipped": skipped, "total": total,
                "selected_types": set(),
                "selected_format": "txt",
                "selected_domain": None,
                "selected_brand": None,
                "selected_country": None,
                "selected_code": None,
                "selected_protocol": None,
                "selected_port": None,
                "selected_network": None,
                "sort_by_domain": False,
                "status": "analysed",
                "user_name": full_name,
                "user_username": uname,
            })
            log_queue(uid, doc.file_name, "queued")

        type_parts = []
        for label, g in (("💳 Cards", cards), ("🔑 Emails", combos), ("📱 Phones", phones),
                         ("🌐 Proxies", proxies), ("🔗 URLs", urls), ("💰 Crypto", cryptos)):
            if g:
                type_parts.append(f"{label}: <b>{len(g)}</b>")
        if skipped:
            type_parts.append(f"🗑️ <b>Skipped:</b> {skipped}")

        await update.message.reply_text(
            f"✅ <b>File analysed!</b>\n"
            f"──────────────────\n"
            f"📁 <code>{doc.file_name}</code>\n\n"
            + "\n".join(type_parts) + "\n\n"
            f"👇 <b>Apply filters & choose format:</b>",
            parse_mode="HTML",
            reply_markup=type_ikb(cards, combos, phones, proxies, urls, cryptos, uid, "txt"),
        )

    except Exception:
        log.exception("handle_document error")
        try:
            await thinking.delete()
        except Exception:
            pass
        await update.message.reply_text(
            "❌ <b>Something went wrong.</b> Please try again.",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )


async def process_queue_item(ctx: ContextTypes.DEFAULT_TYPE, queue_item: dict) -> None:
    uid = queue_item["uid"]
    base = queue_item["base"]
    fmt = queue_item.get("selected_format", "txt")
    selected_types = queue_item.get("selected_types", set())
    domain = queue_item.get("selected_domain", None)
    brand = queue_item.get("selected_brand", None)
    country = queue_item.get("selected_country", None)
    code = queue_item.get("selected_code", None)
    protocol = queue_item.get("selected_protocol", None)
    port = queue_item.get("selected_port", None)
    network = queue_item.get("selected_network", None)
    sort_by_domain = queue_item.get("sort_by_domain", False)
    user_name = queue_item.get("user_name", "")
    user_username = queue_item.get("user_username", "")

    def pick(key):
        return queue_item[key] if (not selected_types or key in selected_types) else []

    out_cards   = pick("cards")
    out_combos  = pick("combos")
    out_phones  = pick("phones")
    out_proxies = pick("proxies")
    out_urls    = pick("urls")
    out_cryptos = pick("cryptos")

    # Apply filters
    if brand and out_cards:
        out_cards = [c for c in out_cards if c["brand"] == brand]
    if country:
        out_cards   = [c for c in out_cards if c.get("country") == country]
        out_combos  = [c for c in out_combos if c.get("country") == country]
        out_phones  = [c for c in out_phones if c.get("country") == country]
        out_proxies = [c for c in out_proxies if c.get("country") == country]
        out_urls    = [c for c in out_urls if c.get("country") == country]
        out_cryptos = [c for c in out_cryptos if c.get("country") == country]
    if code and out_phones:
        out_phones = [c for c in out_phones if c.get("code") == code]
    if protocol and out_proxies:
        out_proxies = [p for p in out_proxies if p.get("protocol") == protocol]
    if port and out_proxies:
        out_proxies = [p for p in out_proxies if p.get("port") == port]
    if network and out_cryptos:
        out_cryptos = [c for c in out_cryptos if c.get("network") == network]
    if domain:
        out_combos, out_urls = filter_by_domain(out_combos, out_urls, domain)
    if sort_by_domain:
        out_combos = sort_combos_by_domain(out_combos)
        out_urls = sort_urls_by_domain(out_urls)

    if not any([out_cards, out_combos, out_phones, out_proxies, out_urls, out_cryptos]):
        await ctx.bot.send_message(
            uid,
            "⚠️ <b>No data left after filtering.</b>\nTry a different type or filter.",
            parse_mode="HTML",
        )
        return

    if fmt == "csv":
        output = build_csv(out_cards, out_combos, out_phones, out_proxies, out_urls, out_cryptos)
        ext = "csv"
    elif fmt == "xlsx":
        output_bytes = build_xlsx(out_cards, out_combos, out_phones, out_proxies, out_urls, out_cryptos)
        ext = "xlsx"
    else:
        output = build_txt(out_cards, out_combos, out_phones, out_proxies, out_urls, out_cryptos)
        ext = "txt"

    if fmt == "xlsx":
        buf = BytesIO(output_bytes)
    else:
        buf = BytesIO(output.encode("utf-8"))
    buf.name = f"{base}_filtered.{ext}"

    parts = [
        (f"💳 Cards: {len(out_cards)}", out_cards),
        (f"🔑 Emails: {len(out_combos)}", out_combos),
        (f"📱 Phones: {len(out_phones)}", out_phones),
        (f"🌐 Proxies: {len(out_proxies)}", out_proxies),
        (f"🔗 URLs: {len(out_urls)}", out_urls),
        (f"💰 Crypto: {len(out_cryptos)}", out_cryptos),
    ]
    lines = [label for label, g in parts if g]
    filters_applied = []
    if brand:
        filters_applied.append(f"💳 {brand}")
    if country:
        filters_applied.append(f"🌍 {country}")
    if code:
        filters_applied.append(f"📞 {code}")
    if protocol:
        filters_applied.append(f"🌐 {protocol}")
    if port:
        filters_applied.append(f"🔌 {port}")
    if network:
        filters_applied.append(f"💰 {network}")
    if domain:
        filters_applied.append(f"📧 {domain}")
    if sort_by_domain:
        filters_applied.append("🔤 sorted")
    if filters_applied:
        lines.append("🧰 " + " · ".join(filters_applied))
    lines.append(f"📄 Format: {ext.upper()}")

    caption = (
        f"✅ <b>Done!</b>\n"
        f"──────────────────\n"
        + "\n".join(lines) + "\n\n"
        + FOOTER
    )

    try:
        await ctx.bot.send_document(
            uid,
            buf,
            filename=f"{base}_filtered.{ext}",
            caption=caption,
            parse_mode="HTML",
            reply_markup=result_ikb(uid),
        )

        upsert_user(uid, user_name, user_username, {
            "cards": len(out_cards), "emails": len(out_combos), "phones": len(out_phones),
            "proxies": len(out_proxies), "urls": len(out_urls), "cryptos": len(out_cryptos),
        })

        log_queue(uid, queue_item["filename"], "done")
        log.info("Processed %s uid=%s cards=%d combos=%d phones=%d proxies=%d urls=%d cryptos=%d fmt=%s",
                 queue_item["filename"], uid, len(out_cards), len(out_combos), len(out_phones),
                 len(out_proxies), len(out_urls), len(out_cryptos), fmt)

    except Exception as e:
        log.error("Error sending file to uid=%s: %s", uid, e)
        log_queue(uid, queue_item["filename"], "error")
        await ctx.bot.send_message(
            uid,
            "❌ <b>Failed to send file.</b> Please try again.",
            parse_mode="HTML",
        )


def _get_user_queue_item(uid: int):
    with queue_lock:
        user_items = [q for q in file_queue if q["uid"] == uid and q["status"] == "analysed"]
        return user_items[-1] if user_items else None


async def _refresh_keyboard(query, item: dict) -> None:
    try:
        await query.message.edit_reply_markup(
            reply_markup=type_ikb(
                item["cards"], item["combos"], item["phones"],
                item["proxies"], item["urls"], item["cryptos"], item["uid"],
                item["selected_format"], item["selected_types"],
                item["selected_domain"], item["sort_by_domain"],
                item["selected_brand"], item["selected_country"], item["selected_code"],
                item["selected_protocol"], item["selected_port"], item["selected_network"],
            )
        )
    except BadRequest:
        pass


def _cycle_filter(query, item, attr, options, label):
    item[attr] = next_choice(item.get(attr), options)
    return f"✅ {label}: {item[attr] or 'All'}"


async def handle_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    uid = query.from_user.id
    data = query.data or ""
    await query.answer()

    if data.startswith("stats:"):
        row = get_user(uid)
        if not row:
            await query.message.reply_text("📊 No stats yet — drop a .txt file to get started!")
        else:
            await query.message.reply_text(user_stats_text(row), parse_mode="HTML")

    elif data == "help":
        await query.message.reply_text(HELP_TEXT, parse_mode="HTML")

    elif data == "owner:stats":
        if uid != OWNER_ID:
            await query.message.reply_text("⛔ Owner only.")
            return
        await query.message.reply_text(global_stats_text(), parse_mode="HTML")

    elif data.startswith("fmt:"):
        parts = data.split(":")
        if len(parts) >= 3:
            fmt = parts[1]
            item = _get_user_queue_item(uid)
            if not item:
                await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
                return
            item["selected_format"] = fmt
            await _refresh_keyboard(query, item)
            await query.answer(f"✅ Format: {fmt.upper()}")

    elif data.startswith("sel:"):
        parts = data.split(":")
        if len(parts) >= 3:
            type_name = parts[1]
            item = _get_user_queue_item(uid)
            if not item:
                await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
                return

            if type_name == "all":
                item["selected_types"] = set()
            elif type_name in item["selected_types"]:
                item["selected_types"].discard(type_name)
            else:
                item["selected_types"].add(type_name)

            await _refresh_keyboard(query, item)
            await query.answer(f"✅ Types: {', '.join(item['selected_types']) or 'All'}")

    elif data.startswith("brand:"):
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
            return
        msg = _cycle_filter(query, item, "selected_brand",
                            get_available_brands(item["cards"]), "Brand")
        await _refresh_keyboard(query, item)
        await query.answer(msg)

    elif data.startswith("cty:"):
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
            return
        msg = _cycle_filter(query, item, "selected_country",
                            get_available_countries(item["cards"], item["combos"], item["phones"],
                                                    item["proxies"], item["urls"], item["cryptos"]),
                            "Country")
        await _refresh_keyboard(query, item)
        await query.answer(msg)

    elif data.startswith("code:"):
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
            return
        msg = _cycle_filter(query, item, "selected_code",
                            get_available_codes(item["phones"]), "Code")
        await _refresh_keyboard(query, item)
        await query.answer(msg)

    elif data.startswith("proto:"):
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
            return
        msg = _cycle_filter(query, item, "selected_protocol",
                            get_available_protocols(item["proxies"]), "Protocol")
        await _refresh_keyboard(query, item)
        await query.answer(msg)

    elif data.startswith("port:"):
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
            return
        msg = _cycle_filter(query, item, "selected_port",
                            get_available_ports(item["proxies"]), "Port")
        await _refresh_keyboard(query, item)
        await query.answer(msg)

    elif data.startswith("net:"):
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
            return
        msg = _cycle_filter(query, item, "selected_network",
                            get_available_networks(item["cryptos"]), "Network")
        await _refresh_keyboard(query, item)
        await query.answer(msg)

    elif data.startswith("seld:"):
        parts = data.split(":")
        if len(parts) >= 3:
            domain = parts[1]
            item = _get_user_queue_item(uid)
            if not item:
                await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
                return
            item["selected_domain"] = domain if item.get("selected_domain") != domain else None
            await _refresh_keyboard(query, item)
            await query.answer(f"✅ Domain: {item['selected_domain'] or 'All'}")

    elif data.startswith("sels:domain:"):
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
            return
        item["sort_by_domain"] = not item.get("sort_by_domain", False)
        await _refresh_keyboard(query, item)
        await query.answer(f"✅ Sort: {'ON' if item['sort_by_domain'] else 'OFF'}")

    elif data.startswith("gen:"):
        parts = data.split(":")
        if len(parts) >= 3:
            item = _get_user_queue_item(uid)
            if not item:
                await query.answer("⏳ No pending file. Send a new file.", show_alert=True)
                return

            item["status"] = "processing"
            await process_queue_item(ctx, item)

            with queue_lock:
                if item in file_queue:
                    file_queue.remove(item)

    elif data.startswith("skip:"):
        # Skip the current file for this user
        item = _get_user_queue_item(uid)
        if not item:
            await query.answer("⏳ No pending file.", show_alert=True)
            return

        with queue_lock:
            if item in file_queue:
                file_queue.remove(item)

        # Delete the message with the keyboard (or edit to show skipped)
        try:
            await query.message.delete()
        except BadRequest:
            pass
        await ctx.bot.send_message(
            uid,
            f"⏭️ <b>File skipped.</b>\n"
            f"📁 <code>{item['filename']}</code> has been discarded.\n"
            f"You can send a new file anytime.",
            parse_mode="HTML",
            reply_markup=MAIN_KB,
        )
        await query.answer("✅ Skipped")


async def handle_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    text = update.message.text or ""
    uid = update.effective_user.id

    if text == "📊 My Stats":
        row = get_user(uid)
        if not row:
            await update.message.reply_text(
                "📊 No stats yet — drop a .txt file to get started!",
                reply_markup=MAIN_KB,
            )
        else:
            await update.message.reply_text(
                user_stats_text(row), parse_mode="HTML", reply_markup=MAIN_KB
            )

    elif text == "ℹ️ Help":
        await update.message.reply_text(HELP_TEXT, parse_mode="HTML", reply_markup=MAIN_KB)

    elif text == "🏷️ About":
        await update.message.reply_text(ABOUT_TEXT, parse_mode="HTML", reply_markup=MAIN_KB)

    elif not text.startswith("/"):
        await update.message.reply_text(
            "👋 <b>Drop a .txt file</b> and I'll clean & filter it instantly.\n\n"
            "Use the buttons below to navigate.",
            parse_mode="HTML", reply_markup=MAIN_KB,
        )


# ── Entry point ─────────────────────────────────────────────────────────────────

def main() -> None:
    init_db()

    if not TOKEN:
        log.error("TELEGRAM_BOT_TOKEN is not set. Export it and restart.")
        raise SystemExit("TELEGRAM_BOT_TOKEN environment variable is required.")

    threading.Thread(target=_start_health, daemon=True).start()

    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("broadcast", cmd_broadcast))
    app.add_handler(CommandHandler("queue", cmd_queue))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(CallbackQueryHandler(handle_callback))

    log.info("Bot starting on port %d", PORT)

    if WEBHOOK_URL:
        log.info("Run mode: WEBHOOK")
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            webhook_url=WEBHOOK_URL,
            drop_pending_updates=True,
        )
    else:
        log.info("Run mode: POLLING (Render-ready — leave WEBHOOK_URL unset)")
        app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()